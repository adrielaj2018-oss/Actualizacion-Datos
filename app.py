# -*- coding: utf-8 -*-
"""
ACTUALIZACION DE DATOS DE TRABAJADORES - PRIZE PRO
Render/GitHub listo. Sin pandas.

Funcionalidades:
- Login administrador/operador con permisos por rol.
- CRUD de usuarios.
- Carga de base trabajadores Excel.
- Captura DNI por digitacion, QR/camara o lector codigo de barras.
- Deteccion automatica al completar 8 digitos + sonido.
- Registro/actualizacion de correo, celular, carnet CONADIS, telefono de emergencia y observacion.
- Persistencia reforzada: SQLite/PostgreSQL + respaldo automatico en Excel maestro.
- Dashboard con KPIs, filtros por fechas, ultimos registros y avance.
- UI responsive, logo PRIZE, sidebar expandible/contraible.
- PWA basica para instalar como app en celular/PC.
"""

import os
import re
import sqlite3
import shutil
from datetime import datetime, date, timedelta
from functools import wraps
from io import BytesIO

from flask import Flask, request, redirect, url_for, session, send_file, render_template_string, flash, jsonify, Response
from openpyxl import Workbook, load_workbook
from werkzeug.security import generate_password_hash, check_password_hash

try:
    import psycopg2
    import psycopg2.extras
except Exception:
    psycopg2 = None

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
PERSIST_DIR = os.getenv("PERSIST_DIR", "/data" if os.path.isdir("/data") else BASE_DIR)
UPLOAD_DIR = os.path.join(PERSIST_DIR, "uploads")
BACKUP_DIR = os.path.join(PERSIST_DIR, "backups")
os.makedirs(UPLOAD_DIR, exist_ok=True)
os.makedirs(BACKUP_DIR, exist_ok=True)

DB_PATH = os.path.join(PERSIST_DIR, "actualizacion_datos.db")
EXCEL_MASTER_PATH = os.path.join(PERSIST_DIR, "DATOS_PERSISTENTES_ACTUALIZACION.xlsx")
DATABASE_URL = os.getenv("DATABASE_URL", "").strip()
USE_POSTGRES = bool(DATABASE_URL)
APP_TZ = os.getenv("APP_TIMEZONE", "America/Lima")

# Render Free no permite Disk persistente para archivos locales.
# La persistencia real debe ser PostgreSQL mediante DATABASE_URL.
def db_backend_name():
    if is_pg():
        return "PostgreSQL"
    return "SQLite local"

def storage_warning():
    if is_pg():
        return "OK: datos principales guardados en PostgreSQL. Excel se genera para descarga/exportación."
    if os.path.abspath(PERSIST_DIR) == os.path.abspath(BASE_DIR):
        return "ADVERTENCIA: SQLite/Excel local no persiste en Render Free al reiniciar o redeployar. Configure DATABASE_URL con PostgreSQL."
    return "Datos guardados en disco persistente /data. Ojo: Render Free Web no permite Disk; use PostgreSQL para plan gratuito."

app = Flask(__name__)
app.secret_key = os.getenv("SECRET_KEY", "cambiar-esta-clave-en-render")
app.config["MAX_CONTENT_LENGTH"] = 25 * 1024 * 1024

# ========================= DB =========================
def is_pg():
    return USE_POSTGRES and psycopg2 is not None

def get_conn():
    if is_pg():
        return psycopg2.connect(DATABASE_URL, cursor_factory=psycopg2.extras.RealDictCursor)
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn

def qmark(sql):
    return sql.replace("?", "%s") if is_pg() else sql

def row_to_dict(row):
    return dict(row) if row else None

def rows_to_dict(rows):
    return [row_to_dict(r) for r in (rows or [])]

def execute(sql, params=(), fetchone=False, fetchall=False, commit=False):
    conn = get_conn(); cur = conn.cursor()
    cur.execute(qmark(sql), params)
    data = None
    if fetchone: data = cur.fetchone()
    if fetchall: data = cur.fetchall()
    if commit: conn.commit()
    cur.close(); conn.close()
    return data

def scalar(sql, params=()):
    r = execute(sql, params, fetchone=True)
    if not r: return 0
    d = row_to_dict(r)
    return list(d.values())[0]

def now_str():
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")

def today_str():
    return date.today().strftime("%Y-%m-%d")

def init_db():
    conn = get_conn(); cur = conn.cursor()
    idtype = "SERIAL PRIMARY KEY" if is_pg() else "INTEGER PRIMARY KEY AUTOINCREMENT"
    cur.execute(qmark(f"""
        CREATE TABLE IF NOT EXISTS usuarios(
            id {idtype}, usuario TEXT UNIQUE NOT NULL, password_hash TEXT NOT NULL,
            nombres TEXT, rol TEXT NOT NULL DEFAULT 'operador', estado TEXT DEFAULT 'ACTIVO', creado_en TEXT
        )
    """))
    cur.execute(qmark(f"""
        CREATE TABLE IF NOT EXISTS trabajadores(
            id {idtype}, empresa TEXT, dni TEXT UNIQUE NOT NULL, trabajador TEXT,
            cargo TEXT, area TEXT, planilla TEXT, estado TEXT DEFAULT 'ACTIVO', fecha_carga TEXT
        )
    """))
    cur.execute(qmark(f"""
        CREATE TABLE IF NOT EXISTS datos_actualizados(
            id {idtype}, dni TEXT UNIQUE NOT NULL, correo TEXT, celular TEXT, observacion TEXT,
            metodo_captura TEXT, nivel_educacion TEXT, procedencia_zona TEXT, indumentaria TEXT, tiempo TEXT, carnet_conadis TEXT, telefono_emergencia TEXT, actualizado_por TEXT, actualizado_en TEXT, fecha TEXT
        )
    """))
    conn.commit()
    # Migraciones suaves. En PostgreSQL un ALTER fallido deja la transacción abortada,
    # por eso cada columna se intenta y se confirma/retrocede por separado.
    def try_add_column(table, col, typ):
        try:
            cur.execute(qmark(f"ALTER TABLE {table} ADD COLUMN {col} {typ}"))
            conn.commit()
        except Exception:
            conn.rollback()
    for col, typ in [("nombres", "TEXT"), ("estado", "TEXT DEFAULT 'ACTIVO'")]:
        try_add_column("usuarios", col, typ)
    try_add_column("datos_actualizados", "fecha", "TEXT")
    for col in ["nivel_educacion", "procedencia_zona", "indumentaria", "tiempo", "carnet_conadis", "telefono_emergencia"]:
        try_add_column("datos_actualizados", col, "TEXT")
    cur.execute(qmark("SELECT id FROM usuarios WHERE usuario=?"), ("admin",))
    if not cur.fetchone():
        cur.execute(qmark("INSERT INTO usuarios(usuario,password_hash,nombres,rol,estado,creado_en) VALUES(?,?,?,?,?,?)"),
                    ("admin", generate_password_hash("admin123"), "ADMINISTRADOR", "admin", "ACTIVO", now_str()))
    conn.commit(); cur.close(); conn.close()

# ========================= UTIL =========================
def normalizar_columna(c):
    c = str(c or "").strip().upper()
    for a,b in {"Á":"A","É":"E","Í":"I","Ó":"O","Ú":"U","Ñ":"N"}.items(): c = c.replace(a,b)
    return re.sub(r"\s+", " ", c)

def limpiar_dni(valor):
    solo = re.sub(r"\D", "", str(valor or ""))
    return solo[-8:] if len(solo) >= 8 else solo

def limpiar_texto(v, upper=True):
    s = "" if v is None else str(v).strip()
    return s.upper() if upper else s

def leer_excel_trabajadores(file_storage):
    wb = load_workbook(file_storage, data_only=True, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows: return [], []
    headers = [normalizar_columna(c) for c in rows[0]]
    data = []
    for row in rows[1:]:
        if not any(row): continue
        item = {}
        for i, h in enumerate(headers):
            if h: item[h] = row[i] if i < len(row) else ""
        data.append(item)
    return headers, data

def excel_mem(headers, rows, sheet):
    wb = Workbook(); ws = wb.active; ws.title = sheet[:31]
    ws.append(headers)
    for row in rows:
        ws.append([row.get(h.lower(), row.get(h, "")) for h in headers])
    for col in ws.columns:
        max_len = max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max(max_len + 2, 12), 42)
    out = BytesIO(); wb.save(out); out.seek(0); return out


def excel_headers_actualizados():
    return ['EMPRESA','DNI','TRABAJADOR','AREA','CARGO','PLANILLA','ESTADO','CORREO','CELULAR','NIVEL_EDUCACION','PROCEDENCIA_ZONA','INDUMENTARIA','TIEMPO','CARNET_CONADIS','TELEFONO_EMERGENCIA','OBSERVACION','METODO_CAPTURA','ACTUALIZADO_POR','ACTUALIZADO_EN']

def ajustar_ancho_excel(ws):
    for col in ws.columns:
        max_len = max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max(max_len + 2, 12), 46)

def crear_backup_excel_maestro(motivo="AUTO"):
    """Crea una copia de seguridad del Excel maestro antes de cambios importantes."""
    try:
        os.makedirs(BACKUP_DIR, exist_ok=True)
        if os.path.exists(EXCEL_MASTER_PATH):
            stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            destino = os.path.join(BACKUP_DIR, f"BACKUP_{motivo}_{stamp}_DATOS_PERSISTENTES_ACTUALIZACION.xlsx")
            shutil.copy2(EXCEL_MASTER_PATH, destino)
            return destino
    except Exception as e:
        print("No se pudo crear backup Excel:", e)
    return ""

def listar_backups(limit=8):
    try:
        os.makedirs(BACKUP_DIR, exist_ok=True)
        archivos = []
        for name in os.listdir(BACKUP_DIR):
            if name.lower().endswith(".xlsx"):
                path = os.path.join(BACKUP_DIR, name)
                archivos.append({
                    "nombre": name,
                    "size_kb": round(os.path.getsize(path)/1024, 1),
                    "fecha": datetime.fromtimestamp(os.path.getmtime(path)).strftime("%Y-%m-%d %H:%M:%S")
                })
        return sorted(archivos, key=lambda x: x["fecha"], reverse=True)[:limit]
    except Exception:
        return []

def sincronizar_excel_maestro():
    try:
        os.makedirs(PERSIST_DIR, exist_ok=True)
        wb = Workbook()
        ws = wb.active
        ws.title = 'TRABAJADORES'
        ws.append(['EMPRESA','DNI','TRABAJADOR','AREA','CARGO','PLANILLA','ESTADO','FECHA_CARGA'])
        trabajadores = rows_to_dict(execute('SELECT empresa,dni,trabajador,area,cargo,planilla,estado,fecha_carga FROM trabajadores ORDER BY trabajador,dni', fetchall=True))
        for r in trabajadores:
            ws.append([r.get('empresa',''), r.get('dni',''), r.get('trabajador',''), r.get('area',''), r.get('cargo',''), r.get('planilla',''), r.get('estado',''), r.get('fecha_carga','')])
        ajustar_ancho_excel(ws)
        ws2 = wb.create_sheet('DATOS_ACTUALIZADOS')
        ws2.append(excel_headers_actualizados())
        actualizados = rows_to_dict(execute('SELECT t.empresa,t.dni,t.trabajador,t.area,t.cargo,t.planilla,t.estado,d.correo,d.celular,d.nivel_educacion,d.procedencia_zona,d.indumentaria,d.tiempo,d.carnet_conadis,d.telefono_emergencia,d.observacion,d.metodo_captura,d.actualizado_por,d.actualizado_en FROM datos_actualizados d JOIN trabajadores t ON t.dni=d.dni ORDER BY d.actualizado_en DESC', fetchall=True))
        for r in actualizados:
            ws2.append([r.get('empresa',''), r.get('dni',''), r.get('trabajador',''), r.get('area',''), r.get('cargo',''), r.get('planilla',''), r.get('estado',''), r.get('correo',''), r.get('celular',''), r.get('nivel_educacion',''), r.get('procedencia_zona',''), r.get('indumentaria',''), r.get('tiempo',''), r.get('carnet_conadis',''), r.get('telefono_emergencia',''), r.get('observacion',''), r.get('metodo_captura',''), r.get('actualizado_por',''), r.get('actualizado_en','')])
        ajustar_ancho_excel(ws2)
        wb.save(EXCEL_MASTER_PATH)
        return True
    except Exception as e:
        print('No se pudo sincronizar Excel maestro:', e)
        return False

def restaurar_desde_excel_si_corresponde():
    if not os.path.exists(EXCEL_MASTER_PATH):
        return
    try:
        wb = load_workbook(EXCEL_MASTER_PATH, data_only=True)
        if scalar('SELECT COUNT(*) AS c FROM trabajadores') == 0 and 'TRABAJADORES' in wb.sheetnames:
            ws = wb['TRABAJADORES']
            headers = [normalizar_columna(c.value) for c in next(ws.iter_rows(min_row=1, max_row=1))]
            conn=get_conn(); cur=conn.cursor(); ins=0
            for row in ws.iter_rows(min_row=2, values_only=True):
                item={headers[i]: row[i] if i < len(row) else '' for i in range(len(headers))}
                dni=limpiar_dni(item.get('DNI'))
                if len(dni)!=8: continue
                cur.execute(qmark('INSERT INTO trabajadores(empresa,dni,trabajador,cargo,area,planilla,estado,fecha_carga) VALUES(?,?,?,?,?,?,?,?)'), (limpiar_texto(item.get('EMPRESA')), dni, limpiar_texto(item.get('TRABAJADOR')), limpiar_texto(item.get('CARGO')), limpiar_texto(item.get('AREA')), limpiar_texto(item.get('PLANILLA')), limpiar_texto(item.get('ESTADO') or 'ACTIVO'), limpiar_texto(item.get('FECHA_CARGA'), upper=False) or now_str()))
                ins += 1
            conn.commit(); cur.close(); conn.close(); print(f'Restaurados trabajadores desde Excel: {ins}')
        if scalar('SELECT COUNT(*) AS c FROM datos_actualizados') == 0 and 'DATOS_ACTUALIZADOS' in wb.sheetnames:
            ws = wb['DATOS_ACTUALIZADOS']
            headers = [normalizar_columna(c.value) for c in next(ws.iter_rows(min_row=1, max_row=1))]
            conn=get_conn(); cur=conn.cursor(); ins=0
            for row in ws.iter_rows(min_row=2, values_only=True):
                item={headers[i]: row[i] if i < len(row) else '' for i in range(len(headers))}
                dni=limpiar_dni(item.get('DNI'))
                if len(dni)!=8: continue
                cur.execute(qmark('SELECT id FROM trabajadores WHERE dni=?'), (dni,))
                if not cur.fetchone(): continue
                cur.execute(qmark('INSERT INTO datos_actualizados(dni,correo,celular,observacion,nivel_educacion,procedencia_zona,indumentaria,tiempo,carnet_conadis,telefono_emergencia,metodo_captura,actualizado_por,actualizado_en,fecha) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?)'), (dni, str(item.get('CORREO') or '').lower(), limpiar_texto(item.get('CELULAR'), upper=False), limpiar_texto(item.get('OBSERVACION')), limpiar_texto(item.get('NIVEL_EDUCACION')), limpiar_texto(item.get('PROCEDENCIA_ZONA')), limpiar_texto(item.get('INDUMENTARIA')), limpiar_texto(item.get('TIEMPO')), limpiar_texto(item.get('CARNET_CONADIS')), limpiar_texto(item.get('TELEFONO_EMERGENCIA'), upper=False), limpiar_texto(item.get('METODO_CAPTURA') or 'EXCEL'), limpiar_texto(item.get('ACTUALIZADO_POR')), limpiar_texto(item.get('ACTUALIZADO_EN'), upper=False) or now_str(), today_str()))
                ins += 1
            conn.commit(); cur.close(); conn.close(); print(f'Restaurados datos actualizados desde Excel: {ins}')
    except Exception as e:
        print('No se pudo restaurar desde Excel maestro:', e)

def login_required(f):
    @wraps(f)
    def wrapper(*args, **kwargs):
        if not session.get("usuario"): return redirect(url_for("login"))
        return f(*args, **kwargs)
    return wrapper

def admin_required(f):
    @wraps(f)
    def wrapper(*args, **kwargs):
        if not session.get("usuario"): return redirect(url_for("login"))
        if session.get("rol") != "admin":
            flash("Solo administrador puede ingresar a esta opcion.", "danger")
            return redirect(url_for("dashboard"))
        return f(*args, **kwargs)
    return wrapper

# ========================= HTML =========================
BASE_HTML = r"""
<!doctype html><html lang="es"><head>
<meta charset="utf-8"><meta name="viewport" content="width=device-width, initial-scale=1">
<meta name="theme-color" content="#16a34a"><link rel="manifest" href="{{ url_for('manifest') }}">
<title>{{ title or 'Actualización de Datos PRIZE' }}</title>
<link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.3/dist/css/bootstrap.min.css" rel="stylesheet">
<link href="https://cdn.jsdelivr.net/npm/bootstrap-icons@1.11.3/font/bootstrap-icons.css" rel="stylesheet">
<script src="https://unpkg.com/html5-qrcode" type="text/javascript"></script>
<script src="https://cdn.jsdelivr.net/npm/chart.js"></script>
<style>
:root{--dark:#07111f;--panel:#0f172a;--panel2:#111827;--green:#16a34a;--green2:#22c55e;--txt:#0f172a;--muted:#64748b;--soft:#eef7f1;--line:#e5e7eb}
*{box-sizing:border-box} body{margin:0;background:linear-gradient(135deg,#f7fafc,#eef7f1);font-family:Inter,Segoe UI,Arial,sans-serif;color:var(--txt)}
.sidebar{position:fixed;left:0;top:0;bottom:0;width:294px;background:radial-gradient(circle at top,#14532d 0,#0f172a 42%,#07111f 100%);color:#fff;padding:18px 14px;z-index:20;transition:.25s;width:294px;box-shadow:12px 0 34px rgba(2,6,23,.20)}
body.collapsed .sidebar{width:86px}.content{margin-left:294px;padding:22px;transition:.25s}body.collapsed .content{margin-left:86px}.logoBox{display:flex;align-items:center;gap:12px;padding:10px 8px 18px;border-bottom:1px solid rgba(255,255,255,.12);margin-bottom:14px}.logoBox img{width:52px;height:52px;border-radius:16px;object-fit:contain;background:white;padding:5px}.brandTitle{font-size:19px;font-weight:900;line-height:1}.brandTitle small{display:block;color:#bbf7d0;font-size:12px;font-weight:700;margin-top:4px}.toggleBtn{border:0;background:rgba(255,255,255,.10);color:white;border-radius:14px;width:42px;height:42px;margin-left:auto}.navLabel{font-size:11px;text-transform:uppercase;letter-spacing:.09em;color:#86efac;margin:16px 10px 6px}.nav-link{display:flex;align-items:center;gap:12px;color:#cbd5e1;border-radius:16px;margin:6px 2px;padding:13px 14px;font-weight:750}.nav-link i{font-size:19px;min-width:24px;text-align:center}.nav-link:hover,.nav-link.active{background:linear-gradient(135deg,var(--green),#15803d);color:white;box-shadow:0 8px 18px rgba(22,163,74,.25)}body.collapsed .navText,body.collapsed .brandTitle,body.collapsed .navLabel,body.collapsed .userBox span{display:none}body.collapsed .logoBox{justify-content:center}body.collapsed .toggleBtn{margin:0 auto}.topbar{background:rgba(255,255,255,.82);backdrop-filter:blur(14px);border:1px solid rgba(226,232,240,.9);border-radius:24px;padding:15px 18px;margin-bottom:18px;box-shadow:0 12px 30px rgba(15,23,42,.06)}.card-pro{border:1px solid rgba(226,232,240,.9);border-radius:24px;box-shadow:0 14px 34px rgba(15,23,42,.07);background:rgba(255,255,255,.96)}.kpi{border-radius:24px;padding:20px;background:white;border:1px solid #e2e8f0;box-shadow:0 12px 26px rgba(15,23,42,.06);position:relative;overflow:hidden}.kpi:after{content:"";position:absolute;right:-18px;top:-18px;width:82px;height:82px;border-radius:50%;background:#dcfce7}.kpiIcon{width:46px;height:46px;border-radius:16px;display:grid;place-items:center;background:#dcfce7;color:#166534;font-size:23px}.btn-pro{border-radius:15px;font-weight:800}.form-control,.form-select{border-radius:15px;padding:11px 13px;border:1px solid #dbe3ef}.form-control:focus,.form-select:focus{box-shadow:0 0 0 .25rem rgba(22,163,74,.16);border-color:#22c55e}.badge-soft{background:#dcfce7;color:#166534;border-radius:999px;padding:8px 12px;font-weight:800}.table{font-size:14px}.table thead th{color:#475569;background:#f8fafc}.userBox{position:absolute;bottom:14px;left:14px;right:14px;background:rgba(255,255,255,.09);border-radius:18px;padding:12px;color:#d1fae5;font-size:13px}.scanPulse{animation:pulse 1s ease-in-out 1}@keyframes pulse{0%{box-shadow:0 0 0 0 rgba(34,197,94,.6)}100%{box-shadow:0 0 0 18px rgba(34,197,94,0)}}
.upper-input{text-transform:uppercase}
body.login-page{min-height:100vh;overflow:hidden;background:linear-gradient(180deg,#f8fbfc 0%,#ecf7f1 52%,#0f7f60 100%);position:relative}
body.login-page:before{content:"";position:fixed;inset:0;z-index:0;background:radial-gradient(circle at 6% 5%,rgba(34,197,94,.16) 0 18%,transparent 19%),radial-gradient(circle at 94% 12%,rgba(20,184,166,.14) 0 22%,transparent 23%),radial-gradient(circle at 22% 42%,rgba(16,185,129,.12),transparent 18%),linear-gradient(180deg,rgba(255,255,255,.78),rgba(255,255,255,.10));pointer-events:none}
body.login-page:after{content:"";position:fixed;left:-4%;right:-4%;bottom:-1px;height:36vh;z-index:0;background:linear-gradient(160deg,rgba(34,197,94,.72),rgba(5,150,105,.88) 46%,rgba(4,78,68,.95));clip-path:polygon(0 42%,18% 28%,39% 18%,61% 34%,79% 22%,100% 35%,100% 100%,0 100%);box-shadow:0 -22px 60px rgba(5,150,105,.22)}
.login-skyline{position:fixed;left:0;right:0;bottom:25vh;height:24vh;z-index:0;opacity:.24;background:linear-gradient(to top,rgba(15,118,110,.45),transparent 80%);clip-path:polygon(0 100%,0 72%,4% 72%,4% 56%,7% 56%,7% 78%,11% 78%,11% 62%,15% 62%,15% 74%,20% 74%,20% 52%,24% 52%,24% 100%,31% 100%,31% 65%,35% 65%,35% 47%,39% 47%,39% 100%,47% 100%,47% 68%,51% 68%,51% 55%,55% 55%,55% 100%,62% 100%,62% 58%,66% 58%,66% 36%,70% 36%,70% 100%,77% 100%,77% 70%,81% 70%,81% 49%,85% 49%,85% 100%,91% 100%,91% 60%,95% 60%,95% 72%,100% 72%,100% 100%)}
body.login-page .content{margin-left:0!important;padding:0;min-height:100vh;position:relative;z-index:1;display:flex;align-items:center;justify-content:center}
.login-wrap{width:100%;min-height:100vh;display:flex;align-items:center;justify-content:center;padding:34px 16px 92px;position:relative}
.login-wrap:before,.login-wrap:after{content:"";position:absolute;width:84px;height:84px;border-radius:999px;background:linear-gradient(135deg,#10b981,#22c55e);box-shadow:0 18px 38px rgba(16,185,129,.22);opacity:.95;display:block}
.login-wrap:before{left:19%;top:46%}.login-wrap:after{right:18%;top:46%}
.login-card{width:min(510px,92vw);padding:78px 54px 36px!important;border-radius:32px!important;background:rgba(255,255,255,.91)!important;backdrop-filter:blur(18px);box-shadow:0 26px 70px rgba(15,23,42,.16)!important;position:relative;overflow:visible}
.login-badge{position:absolute;top:-62px;left:50%;transform:translateX(-50%);width:124px;height:124px;border-radius:999px;background:rgba(255,255,255,.96);display:grid;place-items:center;border:1px solid rgba(226,232,240,.95);box-shadow:0 18px 42px rgba(22,163,74,.18)}
.login-badge i{font-size:54px;color:#15915d}.login-title{font-size:38px;letter-spacing:.5px;color:#1e293b}.login-sub{font-size:16px;color:#64748b}.login-line{width:54px;height:3px;background:#10b981;border-radius:99px;margin:16px auto 20px}.login-field{position:relative}.login-field i{position:absolute;left:14px;top:43px;width:34px;height:34px;border-radius:10px;background:#ecfdf5;color:#15915d;display:grid;place-items:center}.login-field .form-control{height:58px;padding-left:58px;background:rgba(255,255,255,.92);box-shadow:0 8px 18px rgba(15,23,42,.05)}.login-btn{height:58px;border-radius:14px!important;background:linear-gradient(135deg,#16a34a,#05855f)!important;border:0!important;box-shadow:0 14px 28px rgba(5,133,95,.25);font-size:18px}.login-footer{position:fixed;left:0;right:0;bottom:26px;text-align:center;color:white;z-index:2;text-shadow:0 1px 8px rgba(0,0,0,.18)}.login-footer .secure{font-weight:700;margin-bottom:8px}.login-divider{display:flex;align-items:center;gap:18px;color:#64748b;margin-top:28px}.login-divider:before,.login-divider:after{content:"";height:1px;background:#e5e7eb;flex:1}
@media(max-width:920px){
.sidebar{position:sticky;top:0;width:100%!important;min-height:auto;bottom:auto;border-radius:0 0 22px 22px;padding:8px 8px 10px;z-index:50;box-shadow:0 10px 26px rgba(2,6,23,.18)}
.content{margin-left:0!important;padding:10px}.logoBox{padding:4px 4px 7px;margin-bottom:7px;border-bottom:1px solid rgba(255,255,255,.10)}.logoBox img{width:32px;height:32px;border-radius:10px}.brandTitle{display:block!important;font-size:14px}.brandTitle small{font-size:9px}.navMenu{display:flex!important;flex-direction:row!important;flex-wrap:nowrap;overflow-x:auto;overflow-y:hidden;gap:8px;padding:8px;border-radius:18px;background:rgba(2,6,23,.35);white-space:nowrap;-webkit-overflow-scrolling:touch;scroll-snap-type:x proximity}.navMenu::-webkit-scrollbar{height:0}.nav-link{flex:0 0 auto;min-width:max-content;margin:0;padding:9px 13px;border-radius:13px;font-size:12px;gap:6px;background:rgba(255,255,255,.08);border:1px solid rgba(255,255,255,.10);scroll-snap-align:start}.nav-link.active{background:linear-gradient(135deg,#10b981,#15803d)!important}.nav-link i{font-size:15px;min-width:16px}.userBox{display:none}.navText{display:inline!important}.navLabel{display:none}.toggleBtn{display:none}.topbar{border-radius:18px;padding:12px}.topbar h2{font-size:22px}.card-pro{border-radius:18px}.row.g-4{--bs-gutter-y:12px}.input-group .btn{padding-left:12px;padding-right:12px}.correo-dominio-input{max-width:180px;min-width:132px}
body.login-page .content{padding:0}.login-wrap{padding-bottom:78px}.login-wrap:before,.login-wrap:after{display:none}.login-card{padding:72px 22px 26px!important}.login-title{font-size:31px}.login-badge{width:105px;height:105px;top:-52px}.login-badge i{font-size:46px}}
</style></head><body class="{{ ('collapsed ' if session.get('sidebar_collapsed') else '') + ('login-page' if not session.get('usuario') else '') }}">
{% if session.get('usuario') %}
<aside class="sidebar">
  <div class="logoBox">
    <img src="{{ url_for('static', filename='logo_prize.jpeg') }}" onerror="this.style.display='none'">
    <div class="brandTitle">PRIZE PRO<small>Actualización de Datos</small></div>
    <button class="toggleBtn" onclick="toggleSidebar()"><i class="bi bi-list"></i></button>
  </div>
  <div class="navLabel">Operación</div>
  <div class="navMenu nav flex-column">
    <a class="nav-link {% if active=='dashboard' %}active{% endif %}" href="{{ url_for('dashboard') }}"><i class="bi bi-speedometer2"></i><span class="navText">Dashboard</span></a>
    <a class="nav-link {% if active=='captura' %}active{% endif %}" href="{{ url_for('captura') }}"><i class="bi bi-upc-scan"></i><span class="navText">Captura DNI</span></a>
    {% if session.get('rol')=='admin' %}
    <a class="nav-link {% if active=='carga' %}active{% endif %}" href="{{ url_for('cargar_base') }}"><i class="bi bi-file-earmark-excel"></i><span class="navText">Cargar base</span></a>
    <a class="nav-link {% if active=='registros' %}active{% endif %}" href="{{ url_for('registros') }}"><i class="bi bi-table"></i><span class="navText">Registros</span></a>
    <a class="nav-link {% if active=='respaldo' %}active{% endif %}" href="{{ url_for('respaldo') }}"><i class="bi bi-cloud-arrow-down"></i><span class="navText">Respaldo</span></a>
    <div class="navLabel">Administrador</div>
    <a class="nav-link {% if active=='usuarios' %}active{% endif %}" href="{{ url_for('usuarios') }}"><i class="bi bi-people"></i><span class="navText">Usuarios</span></a>
    {% endif %}
    <a class="nav-link" href="{{ url_for('logout') }}"><i class="bi bi-box-arrow-right"></i><span class="navText">Salir</span></a>
  </div>
  <div class="userBox"><i class="bi bi-person-circle me-1"></i><span>{{ session.get('usuario') }} · {{ session.get('rol')|upper }}</span></div>
</aside>
{% endif %}
<main class="content">
{% with messages=get_flashed_messages(with_categories=true) %}{% if messages %}{% for cat,msg in messages %}<div class="alert alert-{{cat}} alert-dismissible fade show card-pro">{{msg}}<button class="btn-close" data-bs-dismiss="alert"></button></div>{% endfor %}{% endif %}{% endwith %}
{{ body|safe }}
</main>
<audio id="sndOk"><source src="data:audio/wav;base64,UklGRjQAAABXQVZFZm10IBAAAAABAAEAESsAACJWAAACABAAZGF0YRAAAAAAAP//AAD//wAA//8AAP//AAD//wAA//8=" type="audio/wav"></audio>
<script src="https://cdn.jsdelivr.net/npm/bootstrap@5.3.3/dist/js/bootstrap.bundle.min.js"></script>
<script>
function toggleSidebar(){document.body.classList.toggle('collapsed'); fetch('/api/sidebar',{method:'POST'}).catch(()=>{});}
function beep(){try{const a=document.getElementById('sndOk'); a.currentTime=0; a.play().catch(()=>{});}catch(e){}}
if('serviceWorker' in navigator){navigator.serviceWorker.register('/sw.js').catch(()=>{});}
</script>
</body></html>
"""

def render_page(body, title="Actualización de Datos PRIZE", active="dashboard", **ctx):
    return render_template_string(BASE_HTML, body=render_template_string(body, **ctx), title=title, active=active)

# ========================= AUTH =========================
@app.route('/login', methods=['GET','POST'])
def login():
    if request.method == 'POST':
        usuario = request.form.get('usuario','').strip()
        password = request.form.get('password','')
        row = row_to_dict(execute('SELECT * FROM usuarios WHERE usuario=?', (usuario,), fetchone=True))
        if row and row.get('estado','ACTIVO') == 'ACTIVO' and check_password_hash(row['password_hash'], password):
            session['usuario']=usuario; session['rol']=row.get('rol','operador'); session['nombres']=row.get('nombres') or usuario
            return redirect(url_for('dashboard'))
        flash('Usuario/clave incorrecta o usuario inactivo.', 'danger')
    body = """
    <div class="login-skyline"></div>
    <div class="login-wrap">
      <div class="card card-pro login-card text-center">
        <div class="login-badge"><i class="bi bi-people-fill"></i></div>
        <h2 class="login-title fw-black fw-bold mb-2">PRIZE PRO</h2>
        <div class="login-sub">Actualización de datos de trabajadores</div>
        <div class="login-line"></div>
        <form method="post" class="text-start">
          <div class="login-field mb-3">
            <label class="form-label fw-bold">Usuario</label>
            <i class="bi bi-person"></i>
            <input class="form-control" name="usuario" required autofocus placeholder="admin">
          </div>
          <div class="login-field mb-3">
            <label class="form-label fw-bold">Clave</label>
            <i class="bi bi-lock"></i>
            <input class="form-control" name="password" type="password" required placeholder="admin123">
          </div>
          <button class="btn btn-success btn-pro login-btn w-100"><i class="bi bi-shield-check me-2"></i>Ingresar</button>
        </form>
        <div class="login-divider">Bienvenido</div>
      </div>
      <div class="login-footer"><div class="secure"><i class="bi bi-shield-check me-2"></i>Sistema seguro y confiable</div><div>© 2025 Prize Pro. Todos los derechos reservados.</div></div>
    </div>"""
    return render_page(body, title='Login', active='')

@app.route('/logout')
def logout():
    session.clear(); return redirect(url_for('login'))

@app.route('/api/sidebar', methods=['POST'])
def api_sidebar():
    session['sidebar_collapsed'] = not bool(session.get('sidebar_collapsed'))
    return jsonify(ok=True)

# ========================= DASHBOARD =========================
@app.route('/')
@login_required
def dashboard():
    desde = request.args.get('desde') or (date.today()-timedelta(days=30)).strftime('%Y-%m-%d')
    hasta = request.args.get('hasta') or today_str()
    total_trab = scalar('SELECT COUNT(*) AS c FROM trabajadores')
    activos = scalar("SELECT COUNT(*) AS c FROM trabajadores WHERE UPPER(COALESCE(estado,''))='ACTIVO'")
    actualizados = scalar('SELECT COUNT(*) AS c FROM datos_actualizados')
    hoy = scalar('SELECT COUNT(*) AS c FROM datos_actualizados WHERE fecha=?', (today_str(),))
    pct = round((actualizados / total_trab * 100), 1) if total_trab else 0
    ultimos = rows_to_dict(execute('''SELECT t.empresa,t.dni,t.trabajador,d.correo,d.celular,d.metodo_captura,d.actualizado_en FROM datos_actualizados d JOIN trabajadores t ON t.dni=d.dni ORDER BY d.actualizado_en DESC LIMIT 8''', fetchall=True))
    por_metodo = rows_to_dict(execute('SELECT metodo_captura AS metodo, COUNT(*) AS total FROM datos_actualizados GROUP BY metodo_captura ORDER BY total DESC', fetchall=True))
    body = """
    <div class="topbar d-flex justify-content-between align-items-center flex-wrap gap-2">
      <div><h2 class="fw-bold mb-0">Dashboard general</h2><div class="text-muted">Control de actualización de correos y celulares · {{ desde }} al {{ hasta }}</div></div>
      <form class="d-flex gap-2 flex-wrap"><input class="form-control" type="date" name="desde" value="{{desde}}"><input class="form-control" type="date" name="hasta" value="{{hasta}}"><button class="btn btn-dark btn-pro">Filtrar</button></form>
    </div>
    <div class="alert alert-{{ 'success' if db_backend == 'PostgreSQL' else 'warning' }} card-pro">
      <b>Persistencia:</b> {{ db_backend }} · {{ storage_msg }}
    </div>
    <div class="row g-3 mb-3">
      <div class="col-md-3"><div class="kpi"><div class="d-flex justify-content-between"><div><div class="text-muted">Trabajadores</div><div class="fs-2 fw-bold">{{total_trab}}</div></div><div class="kpiIcon"><i class="bi bi-people"></i></div></div></div></div>
      <div class="col-md-3"><div class="kpi"><div class="d-flex justify-content-between"><div><div class="text-muted">Activos</div><div class="fs-2 fw-bold">{{activos}}</div></div><div class="kpiIcon"><i class="bi bi-person-check"></i></div></div></div></div>
      <div class="col-md-3"><div class="kpi"><div class="d-flex justify-content-between"><div><div class="text-muted">Actualizados</div><div class="fs-2 fw-bold text-success">{{actualizados}}</div></div><div class="kpiIcon"><i class="bi bi-check2-circle"></i></div></div></div></div>
      <div class="col-md-3"><div class="kpi"><div class="d-flex justify-content-between"><div><div class="text-muted">Hoy</div><div class="fs-2 fw-bold">{{hoy}}</div></div><div class="kpiIcon"><i class="bi bi-calendar2-check"></i></div></div></div></div>
    </div>
    <div class="row g-3 mb-3">
      <div class="col-lg-5"><div class="card card-pro p-4"><h5 class="fw-bold">Avance total</h5><div class="progress" style="height:24px;border-radius:20px"><div class="progress-bar bg-success" style="width:{{pct}}%">{{pct}}%</div></div><div class="text-muted mt-2">{{actualizados}} de {{total_trab}} trabajadores.</div></div></div>
      <div class="col-lg-7"><div class="card card-pro p-4"><h5 class="fw-bold">Métodos de captura</h5><canvas id="chartMetodo" height="80"></canvas></div></div>
    </div>
    <div class="card card-pro p-3"><div class="d-flex justify-content-between align-items-center"><h5 class="fw-bold mb-0">Últimas actualizaciones</h5><a class="btn btn-success btn-pro" href="{{ url_for('captura') }}">Nueva captura</a></div><div class="table-responsive mt-3"><table class="table table-hover align-middle"><thead><tr><th>Empresa</th><th>DNI</th><th>Trabajador</th><th>Correo</th><th>Celular</th><th>Método</th><th>Fecha</th></tr></thead><tbody>{% for r in ultimos %}<tr><td>{{r.empresa}}</td><td class="fw-bold">{{r.dni}}</td><td>{{r.trabajador}}</td><td>{{r.correo}}</td><td>{{r.celular}}</td><td>{{r.metodo_captura}}</td><td>{{r.actualizado_en}}</td></tr>{% else %}<tr><td colspan="7" class="text-center text-muted py-4">Sin registros todavía.</td></tr>{% endfor %}</tbody></table></div></div>
    <script>
    new Chart(document.getElementById('chartMetodo'),{type:'bar',data:{labels:{{ metodos|safe }},datasets:[{label:'Registros',data:{{ valores|safe }} }]},options:{plugins:{legend:{display:false}},scales:{y:{beginAtZero:true,ticks:{precision:0}}}}});
    </script>
    """
    import json
    return render_page(body, active='dashboard', desde=desde, hasta=hasta, total_trab=total_trab, activos=activos, actualizados=actualizados, hoy=hoy, pct=pct, ultimos=ultimos, metodos=json.dumps([x.get('metodo') or 'SIN METODO' for x in por_metodo]), valores=json.dumps([x.get('total') for x in por_metodo]), db_backend=db_backend_name(), storage_msg=storage_warning())

# ========================= CAPTURA =========================
@app.route('/captura')
@login_required
def captura():
    body = """
    <div class="topbar d-flex justify-content-between align-items-center flex-wrap gap-2">
      <div><h2 class="fw-bold mb-0">Captura automática de DNI</h2><div class="text-muted">QR, código de barras, lector USB o digitación manual.</div></div>
      <span class="badge-soft"><i class="bi bi-volume-up me-1"></i>Sonido activo</span>
    </div>
    <div class="row g-4">
      <div class="col-lg-5"><div class="card card-pro p-4" id="scanCard">
        <h5 class="fw-bold"><i class="bi bi-upc-scan me-2"></i>Detector</h5>
        <label class="form-label fw-bold mt-2">DNI / QR / Código de barras</label>
        <div class="input-group mb-3"><input id="dni" class="form-control" placeholder="Escanee o digite DNI" maxlength="30" autofocus autocomplete="off"><button class="btn btn-success btn-pro" onclick="buscarDni('DIGITACION')">Buscar</button></div>
        <div class="d-grid gap-2"><button class="btn btn-outline-success btn-pro" onclick="iniciarCamara()"><i class="bi bi-camera-video me-1"></i>Escanear con cámara</button><button class="btn btn-outline-secondary btn-pro" onclick="detenerCamara()">Detener cámara</button></div>
        <div id="reader" class="mt-3" style="width:100%;display:none"></div>
        <div class="alert alert-light border mt-3 small mb-0"><b>Automático:</b> al detectar 8 dígitos busca solo y emite sonido.</div>
      </div></div>
      <div class="col-lg-7"><div class="card card-pro p-4"><h5 class="fw-bold"><i class="bi bi-person-lines-fill me-2"></i>Datos encontrados</h5><div id="msg" class="alert alert-info">Esperando captura de DNI.</div>
        <form id="frm" style="display:none" onsubmit="guardarDatos(event)"><input type="hidden" id="metodo_captura" value="DIGITACION"><div class="row g-3">
          <div class="col-md-4"><label class="fw-bold">DNI</label><input class="form-control" id="f_dni" readonly></div><div class="col-md-8"><label class="fw-bold">Trabajador</label><input class="form-control" id="trabajador" readonly></div>
          <div class="col-md-6"><label class="fw-bold">Empresa</label><input class="form-control" id="empresa" readonly></div><div class="col-md-6"><label class="fw-bold">Área / Cargo</label><input class="form-control" id="area_cargo" readonly></div>
          <div class="col-md-7"><label class="fw-bold">Correo electrónico</label><div class="input-group"><input class="form-control" id="correo_usuario" placeholder="USUARIO" autocomplete="off"><input class="form-control correo-dominio-input" id="correo_dominio" list="dominiosCorreo" placeholder="@gmail.com" value="@gmail.com" autocomplete="off"><datalist id="dominiosCorreo"><option value="@gmail.com"><option value="@hotmail.com"><option value="@outlook.com"></datalist></div><div class="form-text">Puede elegir o escribir otro dominio desde @, ejemplo: @empresa.com</div><input type="hidden" id="correo"></div><div class="col-md-5"><label class="fw-bold">Celular</label><input class="form-control" id="celular" placeholder="999999999" maxlength="20"></div>
          <div class="col-md-6"><label class="fw-bold">Nivel educación</label><select class="form-select upper-input" id="nivel_educacion"><option value="">SELECCIONE</option><option>PRIMARIA</option><option>SECUNDARIA</option><option>TÉCNICO</option><option>SUPERIOR</option><option>UNIVERSITARIO</option><option>OTRO</option></select></div>
          <div class="col-md-6"><label class="fw-bold">Procedencia (Zona)</label><input class="form-control upper-input" id="procedencia_zona" placeholder="ZONA / PROCEDENCIA"></div>
          <div class="col-md-6"><label class="fw-bold">Indumentaria</label><select class="form-select" id="indumentaria"><option value="">SELECCIONE</option><option>SI</option><option>NO</option></select></div>
          <div class="col-md-6"><label class="fw-bold">Tiempo</label><input class="form-control upper-input" id="tiempo" placeholder="EJEMPLO: 3 MESES / 1 AÑO"></div>
          <div class="col-md-6"><label class="fw-bold">Carnet CONADIS</label><select class="form-select" id="carnet_conadis"><option value="">SELECCIONE</option><option>SI</option><option>NO</option></select></div>
          <div class="col-md-6"><label class="fw-bold">N° telefónico emergencia</label><input class="form-control" id="telefono_emergencia" placeholder="TELÉFONO DE EMERGENCIA" maxlength="20"></div>
          <div class="col-12"><label class="fw-bold">Observación</label><textarea class="form-control upper-input" id="observacion" rows="2" placeholder="OPCIONAL"></textarea></div>
          <div class="col-md-6"><label class="fw-bold">Fecha automática</label><input class="form-control" id="fecha_auto" readonly></div><div class="col-md-6"><label class="fw-bold">Actualizado por</label><input class="form-control" value="{{ session.get('usuario') }}" readonly></div>
        </div><button class="btn btn-success btn-pro mt-3 px-4"><i class="bi bi-save me-1"></i>Guardar actualización</button><button type="button" class="btn btn-outline-secondary btn-pro mt-3" onclick="limpiar()">Nueva captura</button></form>
      </div></div>
    </div>
    <script>
    let html5QrCode=null, buscando=false, timer=null;
    const dniInput=document.getElementById('dni'); dniInput.focus();
    document.addEventListener('input', (e)=>{ if(e.target.classList && e.target.classList.contains('upper-input')) e.target.value=e.target.value.toUpperCase(); });
    document.getElementById('correo_usuario')?.addEventListener('input', e=>{ e.target.value=e.target.value.replace(/@.*/, '').replace(/\s/g,'').toLowerCase(); });
    document.getElementById('correo_dominio')?.addEventListener('input', e=>{ let v=(e.target.value||'').replace(/\s/g,'').toLowerCase(); if(v && !v.startsWith('@')) v='@'+v.replace(/^@+/, ''); e.target.value=v; armarCorreo(); });
    function normalizarDominio(v){ v=(v||'').replace(/\s/g,'').toLowerCase(); if(!v) return '@gmail.com'; if(!v.startsWith('@')) v='@'+v.replace(/^@+/, ''); return v; }
    function armarCorreo(){ const u=(correo_usuario.value||'').replace(/@.*/, '').replace(/\s/g,'').toLowerCase(); const d=normalizarDominio(correo_dominio.value); correo_usuario.value=u; correo_dominio.value=d; correo.value = u ? (u + d) : ''; return correo.value; }
    function cargarCorreo(c){ c=(c||'').trim().toLowerCase(); if(!c){correo_usuario.value=''; correo_dominio.value='@gmail.com'; correo.value=''; return;} const at=c.indexOf('@'); if(at>=0){correo_usuario.value=c.slice(0,at); correo_dominio.value=c.slice(at) || '@gmail.com';} else {correo_usuario.value=c; correo_dominio.value='@gmail.com';} armarCorreo(); }
    dniInput.addEventListener('input',()=>{clearTimeout(timer); const d=limpiarDni(dniInput.value); if(d.length>=8){dniInput.value=d; timer=setTimeout(()=>buscarDni('AUTO/LECTOR'),220);}});
    dniInput.addEventListener('keydown',e=>{if(e.key==='Enter'){e.preventDefault();buscarDni('DIGITACION');}});
    function limpiarDni(v){let raw=(v||'').toString(); let m=raw.match(/(?:^|\D)(\d{8})(?:\D|$)/); let d=m?m[1]:raw.replace(/\D/g,''); return d.length>=8?d.slice(-8):d;}
    function mostrar(txt,tipo){const m=document.getElementById('msg'); m.className='alert alert-'+tipo; m.innerText=txt;}
    async function buscarDni(metodo='DIGITACION'){
      if(buscando) return; const dni=limpiarDni(dniInput.value); dniInput.value=dni; document.getElementById('metodo_captura').value=metodo;
      if(dni.length!==8){mostrar('Ingrese un DNI válido de 8 dígitos.','warning'); return;} buscando=true;
      const r=await fetch('/api/trabajador/'+dni); const data=await r.json(); buscando=false;
      if(!data.ok){document.getElementById('frm').style.display='none'; mostrar(data.msg,'danger'); return;}
      beep(); document.getElementById('scanCard').classList.add('scanPulse'); setTimeout(()=>document.getElementById('scanCard').classList.remove('scanPulse'),900);
      const t=data.trabajador, a=data.actualizacion||{}; document.getElementById('frm').style.display='block';
      f_dni.value=t.dni||''; trabajador.value=t.trabajador||''; empresa.value=t.empresa||''; area_cargo.value=((t.area||'')+' / '+(t.cargo||'')).replace(/^ \/ | \/ $/g,''); cargarCorreo(a.correo||''); celular.value=a.celular||''; nivel_educacion.value=a.nivel_educacion||''; procedencia_zona.value=a.procedencia_zona||''; indumentaria.value=a.indumentaria||''; tiempo.value=a.tiempo||''; carnet_conadis.value=a.carnet_conadis||''; telefono_emergencia.value=a.telefono_emergencia||''; observacion.value=a.observacion||''; fecha_auto.value=new Date().toLocaleString();
      mostrar((a.correo||a.celular)?'Trabajador encontrado. Ya tiene datos, puede modificarlos.':'Trabajador encontrado. Complete correo y celular.','success'); correo.focus();
    }
    async function guardarDatos(e){e.preventDefault(); armarCorreo(); const payload={dni:f_dni.value,correo:correo.value,celular:celular.value,observacion:observacion.value.toUpperCase(),nivel_educacion:nivel_educacion.value.toUpperCase(),procedencia_zona:procedencia_zona.value.toUpperCase(),indumentaria:indumentaria.value.toUpperCase(),tiempo:tiempo.value.toUpperCase(),carnet_conadis:carnet_conadis.value.toUpperCase(),telefono_emergencia:telefono_emergencia.value,metodo_captura:metodo_captura.value}; const r=await fetch('/api/guardar',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify(payload)}); const data=await r.json(); mostrar(data.msg,data.ok?'success':'danger'); if(data.ok){beep(); setTimeout(()=>limpiar('Nuevo registro listo. Escanee o digite otro DNI.'),650);} }
    function limpiar(msj='Esperando captura de DNI.'){document.getElementById('frm').style.display='none'; dniInput.value=''; ['f_dni','trabajador','empresa','area_cargo','correo_usuario','correo','celular','nivel_educacion','procedencia_zona','indumentaria','tiempo','carnet_conadis','telefono_emergencia','observacion','fecha_auto'].forEach(id=>{const el=document.getElementById(id); if(el) el.value='';}); mostrar(msj,'info'); dniInput.focus();}
    function iniciarCamara(){document.getElementById('reader').style.display='block'; html5QrCode=new Html5Qrcode('reader'); html5QrCode.start({facingMode:'environment'},{fps:10,qrbox:250},decoded=>{dniInput.value=limpiarDni(decoded); buscarDni('QR/CAMARA'); detenerCamara();}).catch(()=>mostrar('No se pudo activar cámara. Revise permisos del navegador.','warning'));}
    function detenerCamara(){if(html5QrCode){html5QrCode.stop().catch(()=>{}); html5QrCode=null;} document.getElementById('reader').style.display='none';}
    </script>
    """
    return render_page(body, active='captura')

# ========================= CARGA =========================
@app.route('/cargar-base', methods=['GET','POST'])
@admin_required
def cargar_base():
    if request.method == 'POST':
        f = request.files.get('archivo')
        if not f or not f.filename.lower().endswith(('.xlsx',)):
            flash('Suba un Excel .xlsx válido. No usa pandas para evitar errores en Render.', 'danger'); return redirect(url_for('cargar_base'))
        try: columnas, filas = leer_excel_trabajadores(f)
        except Exception as e:
            flash(f'No se pudo leer Excel: {e}', 'danger'); return redirect(url_for('cargar_base'))
        if 'DNI' not in columnas:
            flash('La plantilla debe tener como mínimo la columna DNI.', 'danger'); return redirect(url_for('cargar_base'))
        crear_backup_excel_maestro('ANTES_CARGA_BASE')
        colmap = {'EMPRESA':'empresa','TRABAJADOR':'trabajador','NOMBRE':'trabajador','APELLIDOS Y NOMBRES':'trabajador','CARGO':'cargo','PUESTO':'cargo','AREA':'area','ÁREA':'area','PLANILLA':'planilla','ESTADO':'estado'}
        ins=upd=omi=0; ahora=now_str(); conn=get_conn(); cur=conn.cursor()
        for r in filas:
            dni=limpiar_dni(r.get('DNI'))
            if len(dni)!=8: omi+=1; continue
            data={'empresa':'','trabajador':'','cargo':'','area':'','planilla':'','estado':'ACTIVO'}
            for col,key in colmap.items():
                if col in columnas:
                    v=limpiar_texto(r.get(col))
                    if v: data[key]=v
            cur.execute(qmark('SELECT id FROM trabajadores WHERE dni=?'),(dni,))
            existe = cur.fetchone()
            if existe:
                cur.execute(qmark('UPDATE trabajadores SET empresa=?,trabajador=?,cargo=?,area=?,planilla=?,estado=?,fecha_carga=? WHERE dni=?'),(data['empresa'],data['trabajador'],data['cargo'],data['area'],data['planilla'],data['estado'],ahora,dni)); upd+=1
            else:
                cur.execute(qmark('INSERT INTO trabajadores(empresa,dni,trabajador,cargo,area,planilla,estado,fecha_carga) VALUES(?,?,?,?,?,?,?,?)'),(data['empresa'],dni,data['trabajador'],data['cargo'],data['area'],data['planilla'],data['estado'],ahora)); ins+=1
        conn.commit(); cur.close(); conn.close()
        sincronizar_excel_maestro()
        flash(f'Carga completada sin borrar registros previos. Insertados: {ins} | Actualizados por DNI: {upd} | Omitidos: {omi}. Se creó backup antes de cargar.', 'success'); return redirect(url_for('cargar_base'))

    q=request.args.get('q','').strip()
    where=''; params=[]
    if q:
        like=f"%{q.upper()}%"
        where='WHERE dni LIKE ? OR UPPER(trabajador) LIKE ? OR UPPER(empresa) LIKE ?'
        params=[like,like,like]
    trabajadores=rows_to_dict(execute(f'SELECT empresa,dni,trabajador,area,cargo,planilla,estado,fecha_carga FROM trabajadores {where} ORDER BY fecha_carga DESC, trabajador LIMIT 300', params, fetchall=True))
    total=scalar('SELECT COUNT(*) AS c FROM trabajadores')
    activos=scalar("SELECT COUNT(*) AS c FROM trabajadores WHERE UPPER(COALESCE(estado,''))='ACTIVO'")
    actualizados=scalar('SELECT COUNT(*) AS c FROM datos_actualizados')
    backups=listar_backups()
    persist_info=PERSIST_DIR
    existe_excel=os.path.exists(EXCEL_MASTER_PATH)
    body = """
    <div class="topbar d-flex justify-content-between align-items-center flex-wrap gap-2">
      <div><h2 class="fw-bold mb-0">Cargar y actualizar base de trabajadores</h2><div class="text-muted">La carga es incremental: si el DNI existe se actualiza, si es nuevo se inserta. No se borra la información anterior.</div></div>
      <div class="d-flex gap-2 flex-wrap"><a class="btn btn-outline-success btn-pro" href="{{ url_for('descargar_base_trabajadores') }}"><i class="bi bi-file-earmark-excel me-1"></i>Descargar base actual</a><a class="btn btn-success btn-pro" href="{{ url_for('descargar_excel_maestro') }}"><i class="bi bi-database-down me-1"></i>Excel maestro</a></div>
    </div>
    <div class="row g-3 mb-3">
      <div class="col-md-4"><div class="kpi"><div class="d-flex justify-content-between"><div><div class="text-muted">Total trabajadores</div><div class="fs-2 fw-bold">{{total}}</div></div><div class="kpiIcon"><i class="bi bi-people"></i></div></div></div></div>
      <div class="col-md-4"><div class="kpi"><div class="d-flex justify-content-between"><div><div class="text-muted">Activos</div><div class="fs-2 fw-bold text-success">{{activos}}</div></div><div class="kpiIcon"><i class="bi bi-person-check"></i></div></div></div></div>
      <div class="col-md-4"><div class="kpi"><div class="d-flex justify-content-between"><div><div class="text-muted">Datos actualizados</div><div class="fs-2 fw-bold">{{actualizados}}</div></div><div class="kpiIcon"><i class="bi bi-check2-circle"></i></div></div></div></div>
    </div>
    <div class="row g-3 mb-3"><div class="col-lg-7"><div class="card card-pro p-4"><h5 class="fw-bold mb-3"><i class="bi bi-upload me-2"></i>Subir Excel para insertar / actualizar</h5><form method="post" enctype="multipart/form-data"><label class="form-label fw-bold">Archivo Excel .xlsx</label><input class="form-control mb-3" type="file" name="archivo" accept=".xlsx" required><button class="btn btn-success btn-pro"><i class="bi bi-upload me-1"></i>Cargar / actualizar base</button> <a class="btn btn-outline-secondary btn-pro" href="{{ url_for('descargar_plantilla') }}">Descargar plantilla</a></form><div class="alert alert-success mt-3 mb-0"><b>Seguro:</b> antes de cada carga se genera un backup automático. La actualización se hace por DNI.</div></div></div><div class="col-lg-5"><div class="card card-pro p-4"><h5 class="fw-bold">Persistencia</h5><p class="text-muted mb-1"><b>Ruta persistente:</b> {{persist_info}}</p><p class="text-muted mb-1"><b>Excel maestro:</b> {{ 'CREADO' if existe_excel else 'PENDIENTE' }}</p><p class="text-muted mb-0">Para Render Free, use <b>DATABASE_URL</b> con PostgreSQL. El Excel es descarga/reporte; no dependa de archivos locales para persistencia.</p></div></div></div>
    <div class="row g-3">
      <div class="col-lg-8"><div class="card card-pro p-3"><div class="d-flex justify-content-between align-items-center flex-wrap gap-2"><h5 class="fw-bold mb-0">Base de trabajadores cargada</h5><form class="d-flex gap-2" method="get"><input class="form-control" name="q" value="{{q}}" placeholder="Buscar DNI, trabajador o empresa"><button class="btn btn-dark btn-pro"><i class="bi bi-search"></i></button></form></div><div class="table-responsive mt-3"><table class="table table-hover align-middle"><thead><tr><th>Empresa</th><th>DNI</th><th>Trabajador</th><th>Área</th><th>Cargo</th><th>Planilla</th><th>Estado</th><th>Fecha carga</th></tr></thead><tbody>{% for r in trabajadores %}<tr><td>{{r.empresa}}</td><td class="fw-bold">{{r.dni}}</td><td>{{r.trabajador}}</td><td>{{r.area}}</td><td>{{r.cargo}}</td><td>{{r.planilla}}</td><td><span class="badge bg-success">{{r.estado}}</span></td><td>{{r.fecha_carga}}</td></tr>{% else %}<tr><td colspan="8" class="text-center text-muted py-4">Todavía no hay base cargada.</td></tr>{% endfor %}</tbody></table></div></div></div>
      <div class="col-lg-4"><div class="card card-pro p-3"><h5 class="fw-bold">Últimos backups</h5><div class="small text-muted mb-2">Se guardan en /data/backups.</div>{% for b in backups %}<div class="border rounded-3 p-2 mb-2"><div class="fw-bold text-truncate">{{b.nombre}}</div><div class="text-muted small">{{b.fecha}} · {{b.size_kb}} KB</div></div>{% else %}<div class="text-muted">Aún no hay backups. Se crearán antes de la próxima carga.</div>{% endfor %}</div></div>
    </div>
    """
    return render_page(body, active='carga', trabajadores=trabajadores, q=q, total=total, activos=activos, actualizados=actualizados, backups=backups, persist_info=persist_info, existe_excel=existe_excel)

@app.route('/base-trabajadores')
@admin_required
def descargar_base_trabajadores():
    rows=rows_to_dict(execute('SELECT empresa,dni,trabajador,area,cargo,planilla,estado,fecha_carga FROM trabajadores ORDER BY trabajador,dni', fetchall=True))
    headers=['EMPRESA','DNI','TRABAJADOR','AREA','CARGO','PLANILLA','ESTADO','FECHA_CARGA']
    data=[{h:r.get(h.lower(),'') for h in headers} for r in rows]
    out=excel_mem(headers,data,'TRABAJADORES')
    return send_file(out, as_attachment=True, download_name=f'base_trabajadores_actual_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

# ========================= REGISTROS =========================
@app.route('/registros')
@admin_required
def registros():
    q=request.args.get('q','').strip(); desde=request.args.get('desde',''); hasta=request.args.get('hasta','')
    where=[]; params=[]
    if q:
        like=f"%{q.upper()}%"; where.append('(t.dni LIKE ? OR UPPER(t.trabajador) LIKE ? OR UPPER(t.empresa) LIKE ?)'); params += [like,like,like]
    if desde: where.append('d.fecha>=?'); params.append(desde)
    if hasta: where.append('d.fecha<=?'); params.append(hasta)
    wh = ('WHERE '+ ' AND '.join(where)) if where else ''
    rows=rows_to_dict(execute(f'''SELECT t.empresa,t.dni,t.trabajador,t.area,t.cargo,d.correo,d.celular,d.nivel_educacion,d.procedencia_zona,d.indumentaria,d.tiempo,d.carnet_conadis,d.telefono_emergencia,d.observacion,d.metodo_captura,d.actualizado_por,d.actualizado_en FROM datos_actualizados d JOIN trabajadores t ON t.dni=d.dni {wh} ORDER BY d.actualizado_en DESC LIMIT 1000''', params, fetchall=True))
    body="""
    <div class="topbar d-flex justify-content-between align-items-center flex-wrap gap-2"><div><h2 class="fw-bold mb-0">Registros actualizados</h2><div class="text-muted">Filtros, seguimiento y exportación.</div></div><div class="d-flex gap-2 flex-wrap"><a class="btn btn-outline-success btn-pro" href="{{ url_for('descargar_excel_maestro') }}"><i class="bi bi-database-down me-1"></i>Excel persistente</a><a class="btn btn-success btn-pro" href="{{ url_for('exportar', q=q, desde=desde, hasta=hasta) }}"><i class="bi bi-file-earmark-excel me-1"></i>Exportar Excel</a></div></div>
    <form class="card card-pro p-3 mb-3" method="get"><div class="row g-2"><div class="col-md-5"><input class="form-control" name="q" value="{{q}}" placeholder="DNI, trabajador o empresa"></div><div class="col-md-3"><input class="form-control" type="date" name="desde" value="{{desde}}"></div><div class="col-md-3"><input class="form-control" type="date" name="hasta" value="{{hasta}}"></div><div class="col-md-1 d-grid"><button class="btn btn-dark btn-pro"><i class="bi bi-search"></i></button></div></div></form>
    <div class="card card-pro p-3"><div class="table-responsive"><table class="table table-hover align-middle"><thead><tr><th>Empresa</th><th>DNI</th><th>Trabajador</th><th>Área</th><th>Cargo</th><th>Correo</th><th>Celular</th><th>Nivel educación</th><th>Procedencia</th><th>Indumentaria</th><th>Tiempo</th><th>Carnet CONADIS</th><th>Tel. emergencia</th><th>Método</th><th>Usuario</th><th>Fecha</th></tr></thead><tbody>{% for r in rows %}<tr><td>{{r.empresa}}</td><td class="fw-bold">{{r.dni}}</td><td>{{r.trabajador}}</td><td>{{r.area}}</td><td>{{r.cargo}}</td><td>{{r.correo}}</td><td>{{r.celular}}</td><td>{{r.nivel_educacion}}</td><td>{{r.procedencia_zona}}</td><td>{{r.indumentaria}}</td><td>{{r.tiempo}}</td><td>{{r.carnet_conadis}}</td><td>{{r.telefono_emergencia}}</td><td>{{r.metodo_captura}}</td><td>{{r.actualizado_por}}</td><td>{{r.actualizado_en}}</td></tr>{% else %}<tr><td colspan="16" class="text-center text-muted py-4">Sin registros.</td></tr>{% endfor %}</tbody></table></div></div>
    """
    return render_page(body, active='registros', rows=rows, q=q, desde=desde, hasta=hasta)

# ========================= USUARIOS =========================
@app.route('/usuarios', methods=['GET','POST'])
@admin_required
def usuarios():
    if request.method=='POST':
        accion=request.form.get('accion'); uid=request.form.get('id'); usuario=request.form.get('usuario','').strip(); nombres=limpiar_texto(request.form.get('nombres')); rol=request.form.get('rol','operador'); estado=request.form.get('estado','ACTIVO'); clave=request.form.get('clave','')
        try:
            if accion=='crear':
                if not usuario or not clave: flash('Usuario y clave son obligatorios.', 'danger')
                else:
                    execute('INSERT INTO usuarios(usuario,password_hash,nombres,rol,estado,creado_en) VALUES(?,?,?,?,?,?)',(usuario,generate_password_hash(clave),nombres,rol,estado,now_str()),commit=True); flash('Usuario creado correctamente.','success')
            elif accion=='editar':
                if clave:
                    execute('UPDATE usuarios SET usuario=?, nombres=?, rol=?, estado=?, password_hash=? WHERE id=?',(usuario,nombres,rol,estado,generate_password_hash(clave),uid),commit=True)
                else:
                    execute('UPDATE usuarios SET usuario=?, nombres=?, rol=?, estado=? WHERE id=?',(usuario,nombres,rol,estado,uid),commit=True)
                flash('Usuario modificado correctamente.','success')
            elif accion=='eliminar':
                if usuario=='admin': flash('No se puede eliminar el usuario admin principal.','danger')
                else: execute('DELETE FROM usuarios WHERE id=?',(uid,),commit=True); flash('Usuario eliminado.','success')
        except Exception as e:
            flash(f'No se pudo procesar: {e}', 'danger')
        return redirect(url_for('usuarios'))
    users=rows_to_dict(execute('SELECT * FROM usuarios ORDER BY id DESC', fetchall=True))
    body="""
    <div class="topbar"><h2 class="fw-bold mb-0">Usuarios del sistema</h2><div class="text-muted">Crear, modificar y eliminar usuarios operadores/administradores.</div></div>
    <div class="row g-3"><div class="col-lg-4"><div class="card card-pro p-4"><h5 class="fw-bold">Crear usuario</h5><form method="post"><input type="hidden" name="accion" value="crear"><label>Usuario</label><input class="form-control mb-2" name="usuario" required><label>Nombres</label><input class="form-control mb-2" name="nombres"><label>Clave</label><input class="form-control mb-2" name="clave" required><label>Rol</label><select class="form-select mb-2" name="rol"><option value="operador">Operador</option><option value="admin">Administrador</option></select><label>Estado</label><select class="form-select mb-3" name="estado"><option>ACTIVO</option><option>INACTIVO</option></select><button class="btn btn-success btn-pro w-100">Crear</button></form></div></div>
    <div class="col-lg-8"><div class="card card-pro p-3"><div class="table-responsive"><table class="table align-middle"><thead><tr><th>Usuario</th><th>Nombres</th><th>Rol</th><th>Estado</th><th>Acción</th></tr></thead><tbody>{% for u in users %}<tr><form method="post"><input type="hidden" name="id" value="{{u.id}}"><td><input class="form-control" name="usuario" value="{{u.usuario}}"></td><td><input class="form-control" name="nombres" value="{{u.nombres or ''}}"></td><td><select class="form-select" name="rol"><option value="operador" {% if u.rol=='operador' %}selected{% endif %}>Operador</option><option value="admin" {% if u.rol=='admin' %}selected{% endif %}>Admin</option></select></td><td><select class="form-select" name="estado"><option {% if u.estado=='ACTIVO' %}selected{% endif %}>ACTIVO</option><option {% if u.estado=='INACTIVO' %}selected{% endif %}>INACTIVO</option></select><input class="form-control mt-1" name="clave" placeholder="Nueva clave opcional"></td><td class="d-flex gap-1"><button class="btn btn-primary btn-sm btn-pro" name="accion" value="editar"><i class="bi bi-pencil"></i></button><button class="btn btn-danger btn-sm btn-pro" name="accion" value="eliminar" onclick="return confirm('¿Eliminar usuario?')"><i class="bi bi-trash"></i></button></td></form></tr>{% endfor %}</tbody></table></div></div></div></div>
    """
    return render_page(body, active='usuarios', users=users)

# ========================= API =========================
@app.route('/api/trabajador/<dni>')
@login_required
def api_trabajador(dni):
    dni=limpiar_dni(dni)
    t=execute('SELECT * FROM trabajadores WHERE dni=?',(dni,),fetchone=True)
    if not t: return jsonify(ok=False,msg='DNI no encontrado en la base de trabajadores.')
    a=execute('SELECT * FROM datos_actualizados WHERE dni=?',(dni,),fetchone=True)
    return jsonify(ok=True, trabajador=row_to_dict(t), actualizacion=row_to_dict(a))

@app.route('/api/guardar', methods=['POST'])
@login_required
def api_guardar():
    data=request.get_json(force=True)
    dni=limpiar_dni(data.get('dni'))
    correo=(data.get('correo') or '').strip().lower()
    celular=re.sub(r'[^0-9+]','',data.get('celular') or '')
    observacion=limpiar_texto(data.get('observacion'))
    nivel_educacion=limpiar_texto(data.get('nivel_educacion'))
    procedencia_zona=limpiar_texto(data.get('procedencia_zona'))
    indumentaria=limpiar_texto(data.get('indumentaria'))
    tiempo=limpiar_texto(data.get('tiempo'))
    carnet_conadis=limpiar_texto(data.get('carnet_conadis'))
    telefono_emergencia=re.sub(r'[^0-9+]','',data.get('telefono_emergencia') or '')
    metodo=limpiar_texto(data.get('metodo_captura') or 'DIGITACION')
    if len(dni)!=8: return jsonify(ok=False,msg='DNI inválido.')
    if correo and not re.match(r'^[^@\s]+@[a-z0-9.-]+\.[a-z]{2,}$', correo): return jsonify(ok=False,msg='Correo inválido. Digite usuario y dominio válido, ejemplo: usuario@gmail.com o usuario@empresa.com.')
    if celular and len(re.sub(r'\D','',celular)) < 7: return jsonify(ok=False,msg='Celular inválido.')
    if indumentaria and indumentaria not in ('SI','NO'): return jsonify(ok=False,msg='Indumentaria debe ser SI o NO.')
    if carnet_conadis and carnet_conadis not in ('SI','NO'): return jsonify(ok=False,msg='Carnet CONADIS debe ser SI o NO.')
    if telefono_emergencia and len(re.sub(r'\D','',telefono_emergencia)) < 7: return jsonify(ok=False,msg='Teléfono de emergencia inválido.')
    if not any([correo, celular, nivel_educacion, procedencia_zona, indumentaria, tiempo, carnet_conadis, telefono_emergencia, observacion]):
        return jsonify(ok=False,msg='Debe registrar al menos un dato.')
    if not execute('SELECT id FROM trabajadores WHERE dni=?',(dni,),fetchone=True): return jsonify(ok=False,msg='El DNI no existe en la base de trabajadores.')
    existe=execute('SELECT id FROM datos_actualizados WHERE dni=?',(dni,),fetchone=True); ahora=now_str(); fecha=today_str(); user=session.get('usuario')
    if existe:
        execute('''UPDATE datos_actualizados SET correo=?,celular=?,observacion=?,nivel_educacion=?,procedencia_zona=?,indumentaria=?,tiempo=?,carnet_conadis=?,telefono_emergencia=?,metodo_captura=?,actualizado_por=?,actualizado_en=?,fecha=? WHERE dni=?''',
                (correo,celular,observacion,nivel_educacion,procedencia_zona,indumentaria,tiempo,carnet_conadis,telefono_emergencia,metodo,user,ahora,fecha,dni),commit=True)
        sincronizar_excel_maestro()
        return jsonify(ok=True,msg='Datos actualizados correctamente y respaldados en Excel.')
    execute('''INSERT INTO datos_actualizados(dni,correo,celular,observacion,nivel_educacion,procedencia_zona,indumentaria,tiempo,carnet_conadis,telefono_emergencia,metodo_captura,actualizado_por,actualizado_en,fecha) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?)''',
            (dni,correo,celular,observacion,nivel_educacion,procedencia_zona,indumentaria,tiempo,carnet_conadis,telefono_emergencia,metodo,user,ahora,fecha),commit=True)
    sincronizar_excel_maestro()
    return jsonify(ok=True,msg='Datos guardados correctamente y respaldados en Excel.')

# ========================= EXPORTS/PWA =========================
@app.route('/exportar')
@login_required
def exportar():
    rows=rows_to_dict(execute('''SELECT t.empresa,t.dni,t.trabajador,t.area,t.cargo,t.planilla,t.estado,d.correo,d.celular,d.nivel_educacion,d.procedencia_zona,d.indumentaria,d.tiempo,d.carnet_conadis,d.telefono_emergencia,d.observacion,d.metodo_captura,d.actualizado_por,d.actualizado_en FROM datos_actualizados d JOIN trabajadores t ON t.dni=d.dni ORDER BY d.actualizado_en DESC''', fetchall=True))
    headers=['EMPRESA','DNI','TRABAJADOR','AREA','CARGO','PLANILLA','ESTADO','CORREO','CELULAR','NIVEL_EDUCACION','PROCEDENCIA_ZONA','INDUMENTARIA','TIEMPO','CARNET_CONADIS','TELEFONO_EMERGENCIA','OBSERVACION','METODO_CAPTURA','ACTUALIZADO_POR','ACTUALIZADO_EN']
    data=[{h:r.get(h.lower(),'') for h in headers} for r in rows]
    out=excel_mem(headers,data,'DATOS_ACTUALIZADOS')
    return send_file(out, as_attachment=True, download_name=f'datos_actualizados_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/plantilla')
@admin_required
def descargar_plantilla():
    out=excel_mem(['EMPRESA','DNI','TRABAJADOR','AREA','CARGO','PLANILLA','ESTADO'], [], 'TRABAJADORES')
    return send_file(out, as_attachment=True, download_name='plantilla_trabajadores.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/excel-maestro')
@admin_required
def descargar_excel_maestro():
    sincronizar_excel_maestro()
    if not os.path.exists(EXCEL_MASTER_PATH):
        flash('No existe Excel maestro todavía.', 'warning')
        return redirect(url_for('dashboard'))
    return send_file(EXCEL_MASTER_PATH, as_attachment=True, download_name='DATOS_PERSISTENTES_ACTUALIZACION.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route('/exportar-todo')
@admin_required
def exportar_todo():
    wb = Workbook()
    ws = wb.active
    ws.title = 'TRABAJADORES'
    ws.append(['EMPRESA','DNI','TRABAJADOR','AREA','CARGO','PLANILLA','ESTADO','FECHA_CARGA'])
    trabajadores = rows_to_dict(execute('SELECT empresa,dni,trabajador,area,cargo,planilla,estado,fecha_carga FROM trabajadores ORDER BY trabajador,dni', fetchall=True))
    for r in trabajadores:
        ws.append([r.get('empresa',''),r.get('dni',''),r.get('trabajador',''),r.get('area',''),r.get('cargo',''),r.get('planilla',''),r.get('estado',''),r.get('fecha_carga','')])
    ajustar_ancho_excel(ws)
    ws2 = wb.create_sheet('DATOS_ACTUALIZADOS')
    ws2.append(excel_headers_actualizados())
    actualizados = rows_to_dict(execute("""SELECT t.empresa,t.dni,t.trabajador,t.area,t.cargo,t.planilla,t.estado,d.correo,d.celular,d.nivel_educacion,d.procedencia_zona,d.indumentaria,d.tiempo,d.carnet_conadis,d.telefono_emergencia,d.observacion,d.metodo_captura,d.actualizado_por,d.actualizado_en FROM datos_actualizados d JOIN trabajadores t ON t.dni=d.dni ORDER BY d.actualizado_en DESC""", fetchall=True))
    for r in actualizados:
        ws2.append([r.get('empresa',''),r.get('dni',''),r.get('trabajador',''),r.get('area',''),r.get('cargo',''),r.get('planilla',''),r.get('estado',''),r.get('correo',''),r.get('celular',''),r.get('nivel_educacion',''),r.get('procedencia_zona',''),r.get('indumentaria',''),r.get('tiempo',''),r.get('carnet_conadis',''),r.get('telefono_emergencia',''),r.get('observacion',''),r.get('metodo_captura',''),r.get('actualizado_por',''),r.get('actualizado_en','')])
    ajustar_ancho_excel(ws2)
    out = BytesIO(); wb.save(out); out.seek(0)
    return send_file(out, as_attachment=True, download_name=f'RESPALDO_TOTAL_ACTUALIZACION_DATOS_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/respaldo')
@admin_required
def respaldo():
    body = """
    <div class="topbar"><h2 class="fw-bold mb-0">Respaldo y persistencia</h2><div class="text-muted">Estado de almacenamiento y descargas de seguridad.</div></div>
    <div class="row g-3">
      <div class="col-lg-6"><div class="card card-pro p-4 h-100"><h5 class="fw-bold"><i class="bi bi-database-check me-2"></i>Motor actual</h5><div class="fs-4 fw-bold mb-2">{{ db_backend }}</div><div class="alert alert-{{ 'success' if db_backend == 'PostgreSQL' else 'warning' }} mb-0">{{ storage_msg }}</div></div></div>
      <div class="col-lg-6"><div class="card card-pro p-4 h-100"><h5 class="fw-bold"><i class="bi bi-download me-2"></i>Descargas</h5><div class="d-grid gap-2"><a class="btn btn-success btn-pro" href="{{ url_for('exportar_todo') }}">Descargar respaldo total Excel</a><a class="btn btn-outline-success btn-pro" href="{{ url_for('descargar_base_trabajadores') }}">Descargar base trabajadores</a><a class="btn btn-outline-success btn-pro" href="{{ url_for('exportar') }}">Descargar datos actualizados</a></div></div></div>
    </div>
    <div class="card card-pro p-4 mt-3"><h5 class="fw-bold">Recomendación para Render Free</h5><p class="mb-1">No uses archivos locales como almacenamiento principal, porque pueden perderse al reiniciar o redesplegar.</p><p class="mb-0">Configura una base PostgreSQL externa o de Render y coloca su URL en la variable <b>DATABASE_URL</b>. Luego descarga respaldos desde esta pantalla.</p></div>
    """
    return render_page(body, active='respaldo', db_backend=db_backend_name(), storage_msg=storage_warning())

@app.route('/manifest.json')
def manifest():
    return jsonify({"name":"PRIZE Actualizacion Datos","short_name":"PRIZE Datos","start_url":"/","display":"standalone","background_color":"#f3f6fb","theme_color":"#16a34a","icons":[{"src":"/static/logo_prize.jpeg","sizes":"192x192","type":"image/jpeg"}]})

@app.route('/sw.js')
def sw():
    return Response("self.addEventListener('install',e=>self.skipWaiting());self.addEventListener('fetch',e=>{});", mimetype='application/javascript')

@app.route('/health')
def health(): return 'OK', 200

init_db()
restaurar_desde_excel_si_corresponde()
sincronizar_excel_maestro()
if __name__ == '__main__':
    port=int(os.getenv('PORT','5000')); app.run(host='0.0.0.0', port=port, debug=True)
